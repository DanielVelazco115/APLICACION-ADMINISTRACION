from openpyxl.styles import PatternFill

def normalizar_nombre(nombre):
    return nombre.strip().lower() if nombre else "sin_depto"

def pintar_encabezados(hoja):
    verde = PatternFill(start_color="CCFFCC", end_color="CCFFCC", fill_type="solid")
    rosa = PatternFill(start_color="FFCCFF", end_color="FFCCFF", fill_type="solid")
    azul = PatternFill(start_color="66CCFF", end_color="66CCFF", fill_type="solid")

    # A1–I1
    for col in range(1, 10):
        hoja.cell(row=1, column=col).fill = verde

    # J1,K1,M1,O1,Q1,R1,T1
    for col in [10,11,13,15,17,18,20]:
        hoja.cell(row=1, column=col).fill = rosa

    # L1,N1,P1,S1
    for col in [12,14,16,19]:
        hoja.cell(row=1, column=col).fill = azul
