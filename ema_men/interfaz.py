import openpyxl
import datetime
import os
from collections import defaultdict
from openpyxl.utils import get_column_letter
from openpyxl.styles import Font, PatternFill
import unicodedata

def normalizar_nombre(nombre):
    if not nombre:
        return "SIN DEPTO"
    nombre = nombre.upper()
    nombre = "".join(
        c for c in unicodedata.normalize("NFD", nombre)
        if unicodedata.category(c) != "Mn"
    )
    return " ".join(nombre.split())

def pintar_encabezados(hoja):
    verde = PatternFill(start_color="CCFFCC", end_color="CCFFCC", fill_type="solid")
    rosa = PatternFill(start_color="FFCCFF", end_color="FFCCFF", fill_type="solid")
    azul = PatternFill(start_color="66CCFF", end_color="66CCFF", fill_type="solid")

    for col in range(1, 10):
        hoja.cell(row=1, column=col).fill = verde
    for col in [10,11,13,15,17,18,20]:
        hoja.cell(row=1, column=col).fill = rosa
    for col in [12,14,16,19]:
        hoja.cell(row=1, column=col).fill = azul

def consolidar(ruta_principal, ruta_nombres, carpeta_salida):
    libro = openpyxl.load_workbook(ruta_principal)
    EMA = libro[libro.sheetnames[0]]

    libro_nombres = openpyxl.load_workbook(ruta_nombres)
    hoja_nombres = libro_nombres[libro_nombres.sheetnames[0]]

    mapa_deptos = {}
    for fila in hoja_nombres.iter_rows(min_row=2, values_only=True):
        nombre, depto = fila[0], fila[1]
        if nombre:
            mapa_deptos[normalizar_nombre(nombre)] = depto or "SIN DEPTO"

    if "Resultado EMA" in libro.sheetnames:
        libro.remove(libro["Resultado EMA"])
    resultado = libro.create_sheet("Resultado EMA")

    encabezado = [cell.value for cell in EMA[1]]
    encabezado.insert(2, "Departamento")
    resultado.append(encabezado)
    pintar_encabezados(resultado)

    grupos = defaultdict(list)
    for fila in EMA.iter_rows(min_row=2, values_only=True):
        if fila[1]:
            nombre = normalizar_nombre(fila[1])
            depto = mapa_deptos.get(nombre) or "SIN DEPTO"
            grupos[depto].append(fila)

    filas_totales = {}
    for depto in sorted(grupos.keys()):
        resultado.append([None, depto])
        inicio = resultado.max_row + 1
        for fila in grupos[depto]:
            nueva = list(fila)
            nueva.insert(2, depto)
            resultado.append(nueva)
        fin = resultado.max_row

        fila_total = [""] * (EMA.max_column + 1)
        fila_total[0] = "TOTAL"
        fila_total[2] = ""
        for col in range(8, 21):  # H–T
            letra = get_column_letter(col)
            fila_total[col-1] = f"=SUM({letra}{inicio}:{letra}{fin})"
        resultado.append(fila_total)
        fila_total_row = resultado.max_row
        filas_totales[depto] = fila_total_row
        resultado.append([])

    # Tablas horizontales lado a lado
    col_inicio = 26
    fila_inicio = 3
    for idx, depto in enumerate(sorted(grupos.keys())):
        base_col = col_inicio + idx*3
        fila_total_row = filas_totales[depto]

        resultado.cell(row=fila_inicio, column=base_col, value=depto).font = Font(bold=True)

        resultado.cell(row=fila_inicio+1, column=base_col, value="IMSS PATRONAL")
        refs_patronal = [f"{c}{fila_total_row}" for c in ["J","K","M","O","Q","R"]]
        resultado.cell(row=fila_inicio+1, column=base_col+1, value=f"=SUM({','.join(refs_patronal)})")

        resultado.cell(row=fila_inicio+2, column=base_col, value="IMSS OBRERO")
        refs_obrero = [f"{c}{fila_total_row}" for c in ["L","N","P","S"]]
        resultado.cell(row=fila_inicio+2, column=base_col+1, value=f"=SUM({','.join(refs_obrero)})")

        resultado.cell(row=fila_inicio+3, column=base_col, value="TOTAL")
        resultado.cell(row=fila_inicio+3, column=base_col+1,
                       value=f"=SUM({get_column_letter(base_col+1)}{fila_inicio+1},{get_column_letter(base_col+1)}{fila_inicio+2})")

    # ============================
    # Tabla resumen final (vertical en Z10)
    # ============================
    fila_resumen = 10
    col_resumen = 26  # Z=26

    resultado.cell(row=fila_resumen, column=col_resumen, value="DEPTO").font = Font(bold=True)
    resultado.cell(row=fila_resumen, column=col_resumen+1, value="PATRONAL").font = Font(bold=True)
    resultado.cell(row=fila_resumen, column=col_resumen+2, value="OBRERO").font = Font(bold=True)
    resultado.cell(row=fila_resumen, column=col_resumen+3, value="TOTAL").font = Font(bold=True)

    for idx, depto in enumerate(sorted(grupos.keys()), start=1):
        base_col = 26 + (idx-1)*3
        fila_resumen_actual = fila_resumen + idx

        resultado.cell(row=fila_resumen_actual, column=col_resumen, value=depto)
        resultado.cell(row=fila_resumen_actual, column=col_resumen+1,
                       value=f"={get_column_letter(base_col+1)}{fila_inicio+1}")
        resultado.cell(row=fila_resumen_actual, column=col_resumen+2,
                       value=f"={get_column_letter(base_col+1)}{fila_inicio+2}")
        resultado.cell(row=fila_resumen_actual, column=col_resumen+3,
                       value=f"={get_column_letter(base_col+1)}{fila_inicio+3}")

    hoy = datetime.date.today()
    mes = hoy.strftime("%B")
    año = hoy.year
    nombre_archivo = f"revision EMA {mes}, {año}.xlsx"
    ruta_salida = os.path.join(carpeta_salida, nombre_archivo)
    libro.save(ruta_salida)
    return ruta_salida

