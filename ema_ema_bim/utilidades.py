import unicodedata
from openpyxl.styles import PatternFill, Font

colores_encabezado = {
    "A1:H1": "FFCCFFCC",
    "I1,J1,L1,N1,P1,Q1,S1": "FFFFCCFF",
    "K1,M1,O1,R1": "FF66CCFF",
    "T1,U1": "FFD9D9D9",
    "V1": "FFDAF2D0",
    "X1,Y1,Z1": "FFCCFFCC",
    "W1,AA1": "FFFBE2D5",
    "AB1": "FF7030A0",
    "AC1": "FF4EA72E",
    "AD1": "FFBE5014"
}

def normalizar_nombre(nombre):
    if not nombre:
        return ""
    nombre = nombre.upper()
    nombre = "".join(
        c for c in unicodedata.normalize("NFD", nombre)
        if unicodedata.category(c) != "Mn"
    )
    palabras = nombre.split()
    palabras.sort()
    return " ".join(palabras)

def pintar_encabezados(hoja):
    for rango, color in colores_encabezado.items():
        if ":" in rango:
            for fila in hoja[rango]:
                for celda in fila:
                    celda.fill = PatternFill(start_color=color, end_color=color, fill_type="solid")
                    celda.font = Font(bold=True, color="000000")
        else:
            for cel in rango.split(","):
                hoja[cel].fill = PatternFill(start_color=color, end_color=color, fill_type="solid")
                hoja[cel].font = Font(bold=True, color="000000")
