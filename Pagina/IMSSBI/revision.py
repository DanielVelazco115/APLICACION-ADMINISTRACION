import openpyxl
from collections import defaultdict
import datetime
import os
from openpyxl.utils import get_column_letter
from openpyxl.styles import Font

# Usamos import relativo para que funcione dentro del paquete Paginas.IMSSBI
from .utilidades import normalizar_nombre, pintar_encabezados

def validar_formula(formula, depto, concepto):
    if not formula or formula.strip() in ["=SUM()", "=", ""]:
        print(f"[ADVERTENCIA] Fórmula inválida en {depto} - {concepto}. Se reemplaza por 0.")
        return "0"
    print(f"[OK] {depto} - {concepto}: {formula}")
    return formula

def construir_refs(columnas, deptos, filas_totales):
    refs = []
    missing = []
    for d in deptos:
        d_norm = normalizar_nombre(d)
        if d_norm in filas_totales:
            fila = filas_totales[d_norm]
            for c in columnas:
                refs.append(f"'Resultado Consolidado'!{c}{fila}")
        else:
            missing.append(d)
    return refs, missing

def fila_concepto_safe(filas_conceptos, depto):
    return filas_conceptos.get(normalizar_nombre(depto), {})

def consolidar(ruta_principal, ruta_nombres, carpeta_salida):
    libro = openpyxl.load_workbook(ruta_principal)

    # Detectar hojas EMA/EBA
    nombre_EMA, nombre_EBA = None, None
    for nombre in libro.sheetnames:
        if "EMA" in nombre.upper():
            nombre_EMA = nombre
        elif "EBA" in nombre.upper():
            nombre_EBA = nombre
            
    if not nombre_EMA or not nombre_EBA:
        raise ValueError("No se encontraron hojas EMA o EBA en el archivo.")

    EMA = libro[nombre_EMA]
    EBA = libro[nombre_EBA]

    # Leer archivo de nombres organizados
    ref_libro = openpyxl.load_workbook(ruta_nombres)
    ref_hoja = ref_libro.active
    mapa_departamentos = {}
    for clave, depto in ref_hoja.iter_rows(min_row=2, max_col=2, values_only=True):
        if clave:
            clave_norm = normalizar_nombre(clave)
            mapa_departamentos[clave_norm] = depto

    # Crear hoja resultado
    if "Resultado Consolidado" in libro.sheetnames:
        libro.remove(libro["Resultado Consolidado"])
    resultado = libro.create_sheet("Resultado Consolidado")

    # Mapas y encabezados
    mapa_indices = {
        "I": 19, "J": 20, "H": 21, "L": 22, "M": 23,
        "N": 24, "O": 25, "P": 26, "K": 28, "Q": 29
    }

    encabezado = [""] * 30
    encabezado[0] = EMA[1][0].value
    encabezado[1] = EMA[1][1].value
    encabezado[2] = "Departamento"
    for idx in range(2, 18):
        encabezado[idx+1] = EMA[1][idx].value
    for col, destino_idx in mapa_indices.items():
        idx_origen = openpyxl.utils.column_index_from_string(col)-1
        encabezado[destino_idx] = EBA[1][idx_origen].value
    encabezado[27] = "Total"
    resultado.append(encabezado)
    pintar_encabezados(resultado)

    # Agrupar filas EMA/EBA
    filas_EMA = [fila for fila in EMA.iter_rows(min_row=2, max_col=19, values_only=True) if fila[0]]
    filas_EBA = [fila for fila in EBA.iter_rows(min_row=2, max_col=18, values_only=True) if fila[0]]

    grupo_EMA = defaultdict(list)
    grupo_EBA = defaultdict(list)
    for fila in filas_EMA:
        grupo_EMA[fila[0]].append(fila)
    for fila in filas_EBA:
        grupo_EBA[fila[0]].append(fila)

    valores_comunes = set(grupo_EMA.keys()) & set(grupo_EBA.keys())

    grupos = defaultdict(list)
    for valor in valores_comunes:
        filas_EMA = grupo_EMA[valor]
        filas_EBA = grupo_EBA[valor]
        max_reps = max(len(filas_EMA), len(filas_EBA))

        clave_norm = normalizar_nombre(filas_EMA[0][1])
        depto = mapa_departamentos.get(clave_norm, "SIN DEPTO") or "SIN DEPTO"

        for i in range(max_reps):
            fila_EMA = list(filas_EMA[i % len(filas_EMA)])
            fila_EBA = filas_EBA[i % len(filas_EBA)]

            fila_final = [""] * 30
            fila_final[0] = fila_EMA[0]
            fila_final[1] = fila_EMA[1]
            fila_final[2] = depto
            for idx in range(2, 18):
                fila_final[idx+1] = fila_EMA[idx]
            for col, destino_idx in mapa_indices.items():
                idx_origen = openpyxl.utils.column_index_from_string(col)-1
                fila_final[destino_idx] = fila_EBA[idx_origen]

            valor_EMA_S = fila_EMA[18] or 0
            valor_EBA_R = fila_EBA[17] or 0
            fila_final[27] = valor_EMA_S + valor_EBA_R

            grupos[depto].append(fila_final)

    # Guardar filas de totales
    filas_totales = {}
    for depto in sorted(grupos.keys()):
        filas = grupos[depto]
        resultado.append([None, depto])
        inicio = resultado.max_row + 1
        for fila in filas:
            resultado.append(fila)
        fin = resultado.max_row

        fila_formula = [""] * 8
        for col in range(9, 31):
            letra = get_column_letter(col)
            formula = f"=SUM({letra}{inicio}:{letra}{fin})"
            fila_formula.append(formula)
        resultado.append(fila_formula)

        fila_total = resultado.max_row
        clave_norm = normalizar_nombre(depto)
        filas_totales[clave_norm] = fila_total

        resultado.append([])
        resultado.append([])
        resultado.append([])

    # Bloque inicial de tablas verticales
    conceptos = ["INFONAVIT","SAR","CESANTÍA PATRONAL","CESANTÍA OBRERO","IMSS PATRONAL","IMSS OBRERO","TOTAL"]
    mapa_resumen = {
        "INFONAVIT": ["W","AA"],
        "SAR": ["V"],
        "CESANTÍA PATRONAL": ["T"],
        "CESANTÍA OBRERO": ["U"],
        "IMSS PATRONAL": ["I","J","L","N","P","Q","S"],
        "IMSS OBRERO": ["K","M","O","R"]
    }

    depto_col_valor = {}
    filas_conceptos = defaultdict(dict)
    col_inicio = 32
    fila_inicio = 2

    for depto in sorted(grupos.keys()):
        depto_norm = normalizar_nombre(depto)
        col_letra1 = get_column_letter(col_inicio)
        col_letra2 = get_column_letter(col_inicio+1)
        depto_col_valor[depto_norm] = col_letra2

        rango = f"{col_letra1}{fila_inicio}:{col_letra2}{fila_inicio}"
        resultado.merge_cells(rango)
        resultado.cell(row=fila_inicio, column=col_inicio, value=depto).font = Font(bold=True)

        for i, concepto in enumerate(conceptos, start=1):
            fila_concepto = fila_inicio + i
            resultado.cell(row=fila_concepto, column=col_inicio, value=concepto)

            if concepto in mapa_resumen:
                cols = mapa_resumen[concepto]
                refs, missing = construir_refs(cols, [depto], filas_totales)
                formula = f"=SUM({','.join(refs)})" if refs else "0"
            elif concepto == "TOTAL":
                formula = f"=SUM({col_letra2}{fila_inicio+1}:{col_letra2}{fila_inicio+len(conceptos)-1})"
            else:
                formula = "0"

            formula = validar_formula(formula, depto, concepto)
            celda_valor = resultado.cell(row=fila_concepto, column=col_inicio+1, value=formula)
            celda_valor.number_format = '"$"#,##0.00'
            filas_conceptos[depto_norm][conceptos[i-1]] = fila_concepto

        # IMSS General
        fila_extra = fila_inicio + len(conceptos) + 1
        resultado.cell(row=fila_extra, column=col_inicio, value="IMSS General")
        fc = filas_conceptos.get(depto_norm, {})
        fila_patronal = fc.get("IMSS PATRONAL")
        fila_obrero = fc.get("IMSS OBRERO")

        if fila_patronal and fila_obrero:
            formula_general = f"='Resultado Consolidado'!{col_letra2}{fila_patronal}+'Resultado Consolidado'!{col_letra2}{fila_obrero}"
        else:
            formula_general = "0"

        resultado.cell(row=fila_extra, column=col_inicio+1, value=validar_formula(formula_general, depto, "IMSS General")).font = Font(bold=True)
        col_inicio += 3

    # Bloque: IMSS Total
    fila_inicio += len(conceptos) + 5
    col_inicio = 32
    for depto in sorted(grupos.keys()):
        depto_norm = normalizar_nombre(depto)
        col_letra1 = get_column_letter(col_inicio)
        col_letra2 = get_column_letter(col_inicio+1)
        
        resultado.cell(row=fila_inicio, column=col_inicio, value=depto).font = Font(bold=True)
        
        # IMSS Total
        resultado.cell(row=fila_inicio+1, column=col_inicio, value="IMSS Total")
        fc = filas_conceptos.get(depto_norm, {})
        f_ces_o = fc.get("CESANTÍA OBRERO")
        f_ims_o = fc.get("IMSS OBRERO")
        f_tot = f"='Resultado Consolidado'!{col_letra2}{f_ces_o}+'Resultado Consolidado'!{col_letra2}{f_ims_o}" if f_ces_o and f_ims_o else "0"
        resultado.cell(row=fila_inicio+1, column=col_inicio+1, value=validar_formula(f_tot, depto, "IMSS Total")).number_format = '"$"#,##0.00'

        # Suma K,M,O,R,U
        resultado.cell(row=fila_inicio+2, column=col_inicio, value="Suma")
        f_idx = filas_totales.get(depto_norm)
        if f_idx:
            refs_c = [f"'Resultado Consolidado'!{c}{f_idx}" for c in ["K","M","O","R","U"]]
            f_sum = f"=SUM({','.join(refs_c)})"
        else:
            f_sum = "0"
        resultado.cell(row=fila_inicio+2, column=col_inicio+1, value=validar_formula(f_sum, depto, "Suma")).number_format = '"$"#,##0.00'

        # Suma anteriores
        resultado.cell(row=fila_inicio+3, column=col_inicio, value="Suma anteriores")
        f_ant = f"=SUM({col_letra2}{fila_inicio+1},{col_letra2}{fila_inicio+2})"
        resultado.cell(row=fila_inicio+3, column=col_inicio+1, value=validar_formula(f_ant, depto, "Suma ant")).number_format = '"$"#,##0.00'
        col_inicio += 3

    # Bloque: Agrupaciones
    fila_inicio += 7
    col_inicio = 32
    agrupaciones = {
        "Administración": ["Administración"],
        "Casco II + Operación": ["Casco II","Operación"],
        "Centro Médico Equino": ["Centro Médico Equino"],
        "Corrales": ["Corrales","Criadero"],
        "Forestal + Jardines": ["Forestal","Jardines"],
        "Producción": ["Producción"],
        "Centro de Reproducción Equina": ["Centro de Reproducción Equina", "Reproducción", "Reproduccion"],
        "Training": ["Training"]
    }
    conceptos_extra = ["INFONAVIT","SAR","CESANTÍA OBRERO","IMSS OBRERO","PATRONAL","SUMA","DIFERENCIA"]

    for nombre, deptos in agrupaciones.items():
        col_letra1 = get_column_letter(col_inicio)
        col_letra2 = get_column_letter(col_inicio+1)
        resultado.cell(row=fila_inicio, column=col_inicio, value=nombre).font = Font(bold=True)

        for i, concepto in enumerate(conceptos_extra, start=1):
            f_conc = fila_inicio + i
            resultado.cell(row=f_conc, column=col_inicio, value=concepto)
            
            if concepto == "INFONAVIT":
                refs, _ = construir_refs(["W","AA"], deptos, filas_totales)
                formula = f"=SUM({','.join(refs)})" if refs else "0"
            elif concepto == "SAR":
                refs, _ = construir_refs(["V"], deptos, filas_totales)
                formula = f"=SUM({','.join(refs)})" if refs else "0"
            elif concepto == "CESANTÍA OBRERO":
                refs, _ = construir_refs(["U"], deptos, filas_totales)
                formula = f"=SUM({','.join(refs)})" if refs else "0"
            elif concepto == "IMSS OBRERO":
                refs, _ = construir_refs(["K","M","O","R"], deptos, filas_totales)
                formula = f"=SUM({','.join(refs)})" if refs else "0"
            elif concepto == "PATRONAL":
                r1, _ = construir_refs(["T"], deptos, filas_totales)
                r2, _ = construir_refs(["I","J","L","N","P","Q","S"], deptos, filas_totales)
                formula = f"=SUM({','.join(r1+r2)})" if (r1+r2) else "0"
            elif concepto == "SUMA":
                formula = f"=SUM({col_letra2}{fila_inicio+1}:{col_letra2}{fila_inicio+5})"
            elif concepto == "DIFERENCIA":
                refs_ind = []
                for d in deptos:
                    dn = normalizar_nombre(d)
                    f_ti = filas_conceptos.get(dn, {}).get("TOTAL")
                    cv = depto_col_valor.get(dn)
                    if f_ti and cv: refs_ind.append(f"'{resultado.title}'!{cv}{f_ti}")
                formula = f"={col_letra2}{fila_inicio+6}-SUM({','.join(refs_ind)})" if refs_ind else "0"
            else:
                formula = "0"

            resultado.cell(row=f_conc, column=col_inicio+1, value=validar_formula(formula, nombre, concepto)).number_format = '"$"#,##0.00'
        col_inicio += 3

    # Guardado final
    hoy = datetime.date.today()
    mes, año = hoy.strftime("%B"), hoy.year
    nombre_archivo = f"revision consolidada {mes}, {año}.xlsx"
    ruta_salida = os.path.join(carpeta_salida, nombre_archivo)
    libro.save(ruta_salida)
    return ruta_salida